import csv
import openpyxl
import sys
from enum import Enum

def clean_positive_employers_en():
    input_file = sys.path[0] + '/positive_employers_en.csv'
    output_file = sys.path[0] + '/../cleaned/2014_12_31.csv'

    # Clear the output file at the beginning
    open(output_file, 'w').close()

    with open(input_file, mode='r', newline='', encoding='utf-8') as infile, \
         open(output_file, mode='w', newline='', encoding='utf-8') as outfile:
        
        reader = csv.reader(infile)
        writer = csv.writer(outfile)

        # Skip the first line
        next(reader)

        # Write the header for the cleansed CSV
        writer.writerow(['Employer', 'Address', 'Positions', 'Province'])

        curr_province = ""

        # Write the cleansed data to the CSV
        for row in reader:
            # check if row is province
            if(row[1] == ""):
                curr_province = row[0]
                continue

            if(row[0].strip() == "Employer" 
               or row[1].strip() == "Address" 
               or (row[2].strip() == "Positions") or (row[2].strip() == "Position Approved")):
                continue

            if len(row) != 3:
                continue

            employer = row[0].strip()
            address = row[1].strip()
            positions = row[2].strip()
            
            writer.writerow([employer, address, positions, curr_province])

def clean_2015():
    input_file = sys.path[0] + '/2015_positive_employers_en.csv'
    output_file = sys.path[0] + '/../cleaned/2015.csv'

    # Clear the output file at the beginning
    open(output_file, 'w').close()

    with open(input_file, mode='r', newline='', encoding='utf-8') as infile, \
         open(output_file, mode='w', newline='', encoding='utf-8') as outfile:
        
        reader = csv.reader(infile)
        writer = csv.writer(outfile)

        # Write the header for the cleansed CSV
        writer.writerow(['Employer', 'Address', 'Positions', 'Province'])

        curr_province = ""

        # Write the cleansed data to the CSV
        for row in reader:
            # check if row is province
            if(row[1] == ""):
                curr_province = row[0]
                continue

            if(row[0].strip() == "Employer" 
               or row[1].strip() == "Address" 
               or (row[2].strip() == "Positions") or (row[2].strip() == "Position Approved")):
                continue

            if len(row) != 3:
                continue

            employer = row[0].strip()
            address = row[1].strip()
            positions = row[2].strip()
            
            writer.writerow([employer, address, positions, curr_province])

def clean_2016():
    input_file = sys.path[0] + '/2016_positive_employer_en.csv'
    output_file = sys.path[0] + '/../cleaned/2016.csv'

    # Clear the output file at the beginning
    open(output_file, 'w').close()

    with open(input_file, mode='r', newline='', encoding='utf-8') as infile, \
         open(output_file, mode='w', newline='', encoding='utf-8') as outfile:
        
        reader = csv.reader(infile)
        writer = csv.writer(outfile)

        # skip the first line
        next(reader)
        # skip the header
        next(reader)

        # Write the header for the cleansed CSV
        writer.writerow(['Province', 'Employer', 'Address', 'Positions', 'NOC 2011', 'NOC Name',])

        curr_province = ""
        same_employer = ""

        # Write the cleansed data to the CSV
        for row in reader:
            # check if row has province
            if(row[0] != ""):
                curr_province = row[0]
            
            if(row[1] != ""):
                same_employer = row[1]

            if len(row) != 5 or row[3] == '':
                continue

            # Province/Territory,Employer,Address,Occupations under NOC 2011,Positions Approved

            if(row[1] == ""):
                employer = same_employer
            else:
                employer = row[1].strip()
            address = row[2].strip()
            occupation = row[3].strip()
            positions = row[4].strip()

            noc_2011 = occupation.split("-")[0]
            noc_name = occupation.split("-")[1]
            writer.writerow([curr_province, employer, address, positions, noc_2011, noc_name])

def clean_2017(input_path, output_path):
    input_file = sys.path[0] + input_path
    output_file = sys.path[0] + output_path

    # Clear the output file at the beginning
    open(output_file, 'w').close()

    with open(input_file, mode='r', newline='', encoding='utf-8') as infile, \
         open(output_file, mode='w', newline='', encoding='utf-8') as outfile:
        
        reader = csv.reader(infile)
        writer = csv.writer(outfile)

        # skip the first line
        next(reader)
        # skip the header
        next(reader)

        # Write the header for the cleansed CSV
        writer.writerow(['Province', 'Employer', 'Address', 'Positions', 'Stream', 'NOC 2011', 'NOC Name',])

        curr_province = ""
        same_employer = ""

        # Write the cleansed data to the CSV
        for row in reader:
            # check if row has province
            if(row[0] != ""):
                curr_province = row[0]
            
            # Province/Territory,Stream,Employer,Address,Occupations under NOC 2011,Positions Approved
            if(row[2] != ""):
                same_employer = row[2]

            if(row[1] != ""):
                same_stream = row[1].strip()

            if len(row) != 6 or row[3] == '':
                continue


            if(row[2] == ""):
                employer = same_employer
            else:
                employer = row[2].strip()
            
            if(row[1] == ""):
                stream = same_stream
            else:
                stream = row[2].strip()

            address = row[3].strip()
            occupation = row[4].strip()
            positions = row[5].strip()

            if(occupation.find("-") != -1):
                noc_2011 = occupation.split("-")[0]
                noc_name = occupation.split("-")[1]
            writer.writerow([curr_province, employer, address, positions, stream, noc_2011, noc_name])

def clean_2021(input_path, output_path):
    input_file = sys.path[0] + input_path
    output_file = sys.path[0] + output_path

    # Clear the output file at the beginning
    open(output_file, 'w').close()

    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active

    with open(output_file, mode='w', newline='', encoding='utf-8') as outfile:
        writer = csv.writer(outfile)

        # Write the header for the cleansed CSV
        writer.writerow(['Province', 'Employer', 'Address', 'Positions', 'Stream', 'NOC 2011', 'NOC Name',])

        curr_province = ""
        same_employer = ""

        # Iterate over the rows in the sheet
        for row in sheet.iter_rows(min_row=3, values_only=True):
            # check if row has province
            if(row[0] != None):
                curr_province = row[0]
            
            # Province/Territory,Stream,Employer,Address,Occupations under NOC 2011,Positions Approved
            if(row[2] != None):
                same_employer = row[2]

            if(row[1] is not None):
                same_stream = row[1].strip()

            if len(row) != 6 or row[3] is None:
                continue


            if(row[2] == ""):
                employer = same_employer
            else:
                employer = row[2].strip()
            
            if(row[1] == ""):
                stream = same_stream
            else:
                stream = row[2].strip()

            address = row[3].strip()
            occupation = row[4].strip()
            positions = row[5]

            if(occupation.find("-") != -1):
                noc_2011 = occupation.split("-")[0]
                noc_name = occupation.split("-")[1]
            writer.writerow([curr_province, employer, address, positions, stream, noc_2011, noc_name])

def clean_2022(input_path, output_path):
    input_file = sys.path[0] + input_path
    output_file = sys.path[0] + output_path

    class Columns(Enum):
        PROVINCE_TERRITORY = 0
        PROGRAM_STREAM = 1
        EMPLOYER = 2
        ADDRESS = 3
        OCCUPATION = 4
        INCORPORATE_STATUS = 5
        APPROVED_LMIAS = 6
        APPROVED_POSITIONS = 7

    # Clear the output file at the beginning
    open(output_file, 'w').close()

    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active

    with open(output_file, mode='w', newline='', encoding='utf-8') as outfile:
        writer = csv.writer(outfile)

        # Write the header for the cleansed CSV
        writer.writerow(['Province', 'Employer', 'Address', 'Positions', 'Stream', 'NOC 2011', 'NOC Name', 'Incorporated Status', 'Approved LMIA'])

        curr_province = ""
        same_employer = ""

        # Province/Territory	Program Stream	Employer 	Address 	Occupation 	Incorporate Status	Approved LMIAs	Approved Positions

        # Iterate over the rows in the sheet
        for row in sheet.iter_rows(min_row=3, values_only=True):
            # check if row has province
            if(row[Columns.PROVINCE_TERRITORY.value] != None):
                curr_province = row[Columns.PROVINCE_TERRITORY.value]

            if(row[Columns.EMPLOYER.value] != None):
                same_employer = row[Columns.EMPLOYER.value]

            if(row[Columns.PROGRAM_STREAM.value] is not None):
                same_stream = row[Columns.PROGRAM_STREAM.value].strip()

            if len(row) != 8 or row[Columns.ADDRESS.value] is None:
                continue


            if(row[Columns.EMPLOYER.value] == ""):
                employer = same_employer
            else:
                employer = row[Columns.EMPLOYER.value].strip()
            
            if(row[Columns.PROGRAM_STREAM.value] == ""):
                stream = same_stream
            else:
                stream = row[Columns.PROGRAM_STREAM.value].strip()

            incorporated_status = row[Columns.INCORPORATE_STATUS.value]
            address = row[Columns.ADDRESS.value].strip()
            occupation = row[Columns.OCCUPATION.value].strip() if row[Columns.OCCUPATION.value] else ""
            positions = row[Columns.APPROVED_POSITIONS.value]
            approved_lmia = row[Columns.APPROVED_LMIAS.value]

            if(occupation.find("-") != -1):
                noc_2011 = occupation.split("-")[0]
                noc_name = occupation.split("-")[1]
            writer.writerow([curr_province, employer, address, positions, stream, noc_2011, noc_name, incorporated_status, approved_lmia])

# clean_positive_employers_en()
# clean_2015()
# clean_2016()
# clean_2017('/2017q1q2_positive_en.csv', '/../cleaned/2017_06_30.csv')
# clean_2017('/2017q3_positive_employer_stream_en.csv', '/../cleaned/2017_09_30.csv')
# clean_2017('/2017q4_positive_employer_en.csv', '/../cleaned/2017_12_31.csv')
# clean_2017('/2018q1_positive_employer_en.csv', '/../cleaned/2018_03_31.csv')
# clean_2017('/2018q2_positive_employer_en.csv', '/../cleaned/2018_06_30.csv')
# clean_2017('/2018q3_positive_en.csv', '/../cleaned/2018_09_30.csv')
# clean_2017('/2018q4_positive_en.csv', '/../cleaned/2018_12_31.csv')

# clean_2017('/tfwp_2019q1_employer_positive_en.csv', '/../cleaned/2019_03.csv')
# clean_2017('/tfwp_2019q2_employer_positive_en.csv', '/../cleaned/2019_06.csv')
# clean_2017('/tfwp_2019q3_positive_en.csv', '/../cleaned/2019_09.csv')
# clean_2017('/tfwp_2019q4_positive_en.csv', '/../cleaned/2019_12.csv')

# clean_2017('/tfwp_2020q1_positive_en.csv', '/../cleaned/2020_03.csv')
# clean_2017('/tfwp_2020q3_positive_en.csv', '/../cleaned/2020_09.csv')

# clean_2021('/TFWP_2021Q2_Positive_EN.xlsx', '/../cleaned/2021_06.csv')
# clean_2021('/TFWP_2021Q3_Positive_EN.xlsx', '/../cleaned/2021_09.csv')

# clean_2022('/tfwp_2022q1_positive_en.xlsx', '/../cleaned/2022_03.csv')
# clean_2022('/tfwp_2022q2_positive_en.xlsx', '/../cleaned/2022_06.csv')
# clean_2022('/tfwp_2022q3_positive_en.xlsx', '/../cleaned/2022_09.csv')
# clean_2022('/tfwp_2022q4_pos_en.xlsx', '/../cleaned/2022_12.csv')

# clean_2022('/tfwp_2023q1_pos_en.xlsx', '/../cleaned/2023_03.csv')
# clean_2022('/tfwp_2023q2_pos_en.xlsx', '/../cleaned/2023_06.csv')
# clean_2022('/tfwp_2023q3_pos_en.xlsx', '/../cleaned/2023_09.csv')
# clean_2022('/tfwp_2023q4_pos_en.xlsx', '/../cleaned/2023_12.csv')
clean_2022('/tfwp_2024q1_pos_en.xlsx', '/../cleaned/2024_03.csv')